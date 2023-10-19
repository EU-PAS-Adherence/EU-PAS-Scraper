from difflib import get_close_matches, SequenceMatcher
from itertools import groupby
import json
import logging
from pathlib import Path
import re
import unicodedata

import cleanco
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
import numpy as np
from scrapy.commands import ScrapyCommand
from scrapy.exceptions import UsageError
from sklearn.cluster import AffinityPropagation
from sklearn.feature_extraction.text import ENGLISH_STOP_WORDS

from eupas.items import Study


class Command(ScrapyCommand):

    requires_project = True

    # and, &, Dept, Department, co, company, R&D
    # delete all chars in () or []
    junk_words = frozenset({
        'pharma', 'pharmaceuticals', 'therapeutics', 'international', 'group',
        'cro', 'kk', 'pvt', 'nhs foundation trust'
    })

    def add_options(self, parser):
        ScrapyCommand.add_options(self, parser)
        group = parser.add_argument_group(title="Custom Grouping Options")
        group.add_argument(
            "-i",
            "--input",
            metavar="FILE",
            default=None,
            help="path to the data"
        )
        group.add_argument(
            "-o",
            "--output",
            metavar="FOLDER",
            default=None,
            help="path to the output folder"
        )
        group.add_argument(
            "-c",
            "--cutoff",
            metavar="CUTOFF",
            default=0.6,
            help="cutoff value for grouping"
        )

    def process_options(self, args, opts):
        ScrapyCommand.process_options(self, args, opts)
        self.input_path = Path(opts.input or "")
        if not self.input_path.is_file():
            raise UsageError(
                "Invalid -i value, use a valid path to a file", print_help=False)
        if self.input_path.suffix != '.json':
            raise UsageError(
                "Invalid -i value, json file expected", print_help=False)

        self.output_folder = Path(opts.output or "")
        if not self.output_folder.is_dir():
            raise UsageError(
                "Invalid -i value, use a valid path to a folder", print_help=False)
        self.output_folder.mkdir(parents=True, exist_ok=True)

        try:
            self.cutoff = float(opts.cutoff)
            assert self.cutoff >= 0 and self.cutoff <= 1
        except (ValueError, AssertionError) as e:
            raise UsageError(
                "Invalid -c value, use a valid float between 0 and 1", print_help=False) from e

    def syntax(self):
        return "field_names [options]"

    def short_desc(self):
        return "Group extracted row value"

    def serialize(self, s, casefold=True):
        def filter_multiple_spaces(s):
            return re.sub(r'\s+', ' ', s)

        def filter_junk_chars(s):
            return re.sub(r'[^\w ]', '', s)

        def sub_junk_words(m):
            return '' if m.group() in self.junk_words.union(ENGLISH_STOP_WORDS) else m.group()

        def filter_junk_words(s):
            return cleanco.basename(re.sub(r'\w+', sub_junk_words, s), suffix=True, prefix=True, middle=True)

        return filter_multiple_spaces(
            filter_junk_words(
                filter_junk_chars(
                    unicodedata.normalize('NFKD', str(s).casefold() if casefold else str(s)).encode(
                        'ASCII', 'ignore').decode('UTF-8', 'ignore')
                ))).strip()

    def affinity_cluster_with_difflib(self, field_name):
        values = sorted(
            list({study[field_name] for study in self.studies if field_name in study}))
        clean_values = [self.serialize(s) for s in values]
        matcher = SequenceMatcher()

        similarity = np.identity(len(values))
        for i in range(1, len(values)):
            for j in range(i):
                matcher.set_seqs(clean_values[i], clean_values[j])
                similarity[i,j] = similarity[j,i] = matcher.ratio()

        clusters = AffinityPropagation(
            affinity='precomputed', max_iter=100, convergence_iter=8).fit_predict(similarity)
        return {int(k): [h[1] for h in list(g)] for k, g in groupby(sorted(zip(clusters, clean_values),
                key=lambda x: x[0]), key=lambda x: x[0])}

    def simple_cluster_with_difflib(self, field_name):
        '''
        Groups field_name using difflib get_close_matches.
        This grouping algorithm isn't optimal.
        It doesn't compare new values with already grouped values.
        '''

        class firstShowTuple(tuple):
            '''
            This class is utilized to pass the serialized value to difflib,
            while storing the original value in the tuple for later retrieval
            '''

            def __str__(self) -> str:
                return self.__getitem__(0).__str__()

        groups = {}

        values = {
            firstShowTuple(
                (self.serialize(study[field_name]), str(study[field_name])))
            for study in self.studies
            if field_name in study
        }

        while len(values) > 1:
            val = values.pop()
            matches = get_close_matches(
                val, values, n=len(values), cutoff=self.cutoff)
            values.difference_update(matches)
            matches.append(val)
            groups.setdefault(self.serialize(
                val[1], casefold=False), sorted([m[1] for m in matches]))

        if len(values) == 1:
            val = values.pop()
            groups.setdefault(
                self.serialize(val[1], casefold=False), [val[1]])

        return groups

    def run(self, args, opts):
        if len(args) == 0:
            raise UsageError(
                "running 'scrapy group' without additional arguments is not supported"
            )

        if not set(args).issubset(set(Study.fields)):
            raise UsageError(
                "At least one -g value isn't a valid field name", print_help=False)

        with self.input_path.open('r') as f:
            self.studies = json.load(f)

        workbook = Workbook()
        workbook.remove(workbook.active)

        logger = logging.getLogger()

        for field_name in args:
            logger.info(f'Grouping field: {field_name}')

            # DEBUG: export unique values
            # values = sorted(list({str(study[field_name])
            #                       for study in self.studies
            #                       if field_name in study
            #                        }))
            # with open(f'{self.output_folder}/values_{field_name}.json', 'w') as f:
            #     json.dump(values, f, indent='\t')

            groups = self.affinity_cluster_with_difflib(field_name)
            # groups = self.simple_cluster_with_difflib(field_name)

            with open(f'{self.output_folder}/{field_name}.json', 'w') as f:
                json.dump(groups, f, indent='\t', sort_keys=True)

            sheet = workbook.create_sheet(field_name)
            for group_name, field_names in sorted(groups.items()):
                group_cell = WriteOnlyCell(value=group_name, ws=sheet)
                field_cells = [WriteOnlyCell(value=name, ws=sheet)
                               for name in field_names]
                sheet.append(
                    [group_cell, field_cells[0] if field_cells else ''])
                if len(field_cells) > 1:
                    start_row = sheet.max_row
                    for field_cell in field_cells[1:]:
                        sheet.append(['', field_cell])
                    sheet.merge_cells(
                        start_row=start_row, end_row=sheet.max_row, start_column=1, end_column=1)

        workbook.save(f'{self.output_folder}/groups.xlsx')
